import tkinter as tk
from tkinter import *
from tkinter import ttk 
import tkinter.messagebox as messagebox
import sv_ttk
import sqlite3
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import csv

def backup_database():
    # Connect to SQLite database
    db_file = 'inventory.db'
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Ask user to select a file for export
    export_file_path = filedialog.asksaveasfilename(defaultextension=".bak", initialfile="backup.bak", filetypes=[("Backup files", "*.bak")])

    if export_file_path:
        try:
            # Execute query to gather required data
            cursor.execute('''SELECT Inventory.Description, Vendors.VendorName, Vendors.RepName, Vendors.RepPhone, Vendors.Discontinued,
                              Location.Location, Location.SubLocation, Location.Discontinued,
                              Inventory.Quantity, Inventory.ReorderLevel, Inventory.Cost, Inventory.Sell, Inventory.Discontinued
                              FROM Inventory
                              INNER JOIN Vendors ON Inventory.VendorID = Vendors.VendorID
                              INNER JOIN Location ON Inventory.LocationID = Location.LocationID''')
            data = cursor.fetchall()

            # Write data to CSV file
            with open(export_file_path, 'w', newline='') as csv_file:
                csv_writer = csv.writer(csv_file)
                # Write header
                csv_writer.writerow(['Description', 'VendorName', 'RepName', 'RepPhone', 'VendorDiscontinued', 
                                     'Location', 'SubLocation', 'LocationDiscontinued', 
                                     'Quantity', 'ReorderLevel', 'Cost', 'Sell', 'InventoryDiscontinued'])
                # Write data
                csv_writer.writerows(data)

            messagebox.showinfo("Success", f"Database records successfully backed up to: {export_file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during backup: {str(e)}")

    # Close database connection
    conn.close()

def restore_database():
    # Connect to SQLite database
    db_file = 'inventory.db'
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Ask user to select the CSV file for import
    import_file_path = filedialog.askopenfilename(filetypes=[("Backup files", "*.bak")])

    if import_file_path:
        try:
            # Read data from CSV file
            with open(import_file_path, 'r', newline='') as csv_file:
                csv_reader = csv.reader(csv_file)
                next(csv_reader)  # Skip header

                for row in csv_reader:
                    # Extract data from each row
                    description, vendor_name, rep_name, rep_phone, vendor_discontinued, location, sub_location, location_discontinued, \
                    quantity, reorder_level, cost, sell, inventory_discontinued = row

                    # Check if the vendor already exists in the database
                    cursor.execute("SELECT VendorID FROM Vendors WHERE VendorName=?", (vendor_name,))
                    vendor_id = cursor.fetchone()
                    if vendor_id:
                        vendor_id = vendor_id[0]
                    else:
                        # Insert new vendor into Vendors table
                        cursor.execute("INSERT INTO Vendors (VendorName, RepName, RepPhone, Discontinued) VALUES (?, ?, ?, ?)",
                                       (vendor_name, rep_name, rep_phone, vendor_discontinued))
                        vendor_id = cursor.lastrowid

                    # Check if the location already exists in the database
                    cursor.execute("SELECT LocationID FROM Location WHERE Location=? AND SubLocation=?", (location,sub_location,))
                    location_id = cursor.fetchone()
                    if location_id:
                        location_id = location_id[0]
                    else:
                        # Insert new location into Location table
                        cursor.execute("INSERT INTO Location (Location, SubLocation, Discontinued) VALUES (?, ?, ?)",
                                       (location, sub_location, location_discontinued))
                        location_id = cursor.lastrowid

                    # Insert inventory data into Inventory table
                    cursor.execute('''INSERT INTO Inventory (Description, VendorID, LocationID, Quantity, ReorderLevel, Cost, Sell, Discontinued)
                                      VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                                   (description, vendor_id, location_id, quantity, reorder_level, cost, sell, inventory_discontinued))

            messagebox.showinfo("Success", f"Data successfully restored from: {import_file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during restore: {str(e)}")
            # Rollback changes if any error occurs
            conn.rollback()
 
     # Commit changes and close database connection
    conn.commit()
    conn.close()
    populate_treeview()

def create_import_file():
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile="import_template.csv", filetypes=[("CSV files", "*.csv")])

    if file_path:
        # Define header rows
        headers = ['Description', 'Vendor', 'RepName', 'RepPhone', 'Location', 'SubLocation', 'Quantity', 'ReorderLevel', 'Cost', 'Sell']

        # Write headers to import.csv
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
        messagebox.showinfo("Success", f"Import template file created at: {file_path}")

def import_records():
    # Ask user to select file for opening
    file_path = filedialog.askopenfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    if not file_path:
        return

    # Connect to the SQLite database
    conn = sqlite3.connect('inventory.db')
    cursor = conn.cursor()

    # Read data from CSV file and populate Vendors, Locations, and Items tables
    with open(file_path, 'r', newline='') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            description = row['Description']
            vendor_name = row['Vendor']
            repname = row['RepName']
            repphone= row['RepPhone']
            location_name = row['Location']
            sub_location = row['SubLocation'] 
            quantity = int(row['Quantity'].replace(',', '').strip())  
            reorder_level = int(row['ReorderLevel'].replace(',', '').strip())  
            cost = int(float(row['Cost'].replace(',', '').replace('$', '')) * 100)
            sell = int(float(row['Sell'].replace(',', '').replace('$', '')) * 100)

            # Check if vendor already exists
            cursor.execute("SELECT VendorID FROM Vendors WHERE VendorName=?", (vendor_name,))
            vendor_id = cursor.fetchone()
            if vendor_id is None:
                cursor.execute("INSERT INTO Vendors (VendorName, RepName, RepPhone, Discontinued) VALUES (?, ?, ?, 'N')", (vendor_name,repname,repphone,))
                vendor_id = cursor.lastrowid
            else:
                vendor_id = vendor_id[0]

            # Check if location already exists
            cursor.execute("SELECT LocationID FROM Location WHERE Location=? AND SubLocation=?", (location_name, sub_location))
            location_id = cursor.fetchone()
            if location_id is None:
                cursor.execute("INSERT INTO Location (Location, SubLocation, Discontinued) VALUES (?, ?, 'N')", (location_name, sub_location))
                location_id = cursor.lastrowid
            else:
                location_id = location_id[0]

            # Insert item into Items table
            cursor.execute("INSERT INTO Inventory (Description, VendorID, LocationID, Quantity, ReorderLevel, Cost, Sell, Discontinued) VALUES (?, ?, ?, ?, ?, ?, ?, 'N')",
                           (description, vendor_id, location_id, quantity, reorder_level, cost, sell))

    # Commit changes and close connection
    conn.commit()
    conn.close()
    populate_treeview()

def check_database():
    db_file = 'inventory.db'

    # Check if the database file exists
    if not os.path.exists(db_file):
        # Connect to SQLite database (or create it if it doesn't exist)
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()

        # Create Vendors table
        cursor.execute('''CREATE TABLE IF NOT EXISTS Vendors (
                        VendorID INTEGER UNIQUE PRIMARY KEY AUTOINCREMENT,
                        VendorName TEXT NOT NULL,
                        RepName TEXT,
                        RepPhone TEXT,
                        Discontinued INTEGER
                        )''')

        # Create Location table
        cursor.execute('''CREATE TABLE IF NOT EXISTS Location (
                        LocationID INTEGER UNIQUE PRIMARY KEY AUTOINCREMENT,
                        Location TEXT NOT NULL,
                        SubLocation TEXT, 
                        Discontinued INTEGER
                        )''')

        # Create Inventory table
        cursor.execute('''CREATE TABLE IF NOT EXISTS Inventory (
                        ItemID INTEGER UNIQUE PRIMARY KEY AUTOINCREMENT,
                        Description TEXT NOT NULL,
                        VendorID INTEGER,
                        LocationID INTEGER,
                        Quantity INTEGER,
                        ReorderLevel INTEGER,
                        Cost REAL,
                        Sell REAL,
                        Discontinued INTEGER,
                        FOREIGN KEY (VendorID) REFERENCES Vendors(VendorID),
                        FOREIGN KEY (LocationID) REFERENCES Location(LocationID)
                        )''')

        # Create Settings table
        cursor.execute('''CREATE TABLE IF NOT EXISTS Settings (
                        Purpose TEXT NOT NULL,
                        DiscName1 TEXT,
                        Discount1 INTEGER,
                        DiscName2 TEXT,
                        Discount2 INTEGER,
                        DiscName3 TEXT,
                        Discount3 INTEGER
                        )''')

        default_values = ("Discounts", "10% Off", 10, "15% Off", 15, "20% Off", 20)

        cursor.execute('''INSERT INTO Settings (Purpose, DiscName1, Discount1, DiscName2, Discount2, DiscName3, Discount3) VALUES (?, ?, ?, ?, ?, ?, ?)''', default_values)

        # Commit changes and close connection
        conn.commit()
        conn.close()

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # Calculate the position of the window
    x = max((screen_width - width) // 2, 0)
    y = max((screen_height - height) // 2 - 25, 0)  # Adjusting for window decorations

    # Set the geometry of the window to center it on the screen
    window.geometry(f"{width}x{height}+{x}+{y}")
    window.minsize(width, height)

def get_mapped_value(input_value):

    value_mapping = {
    "Product Description": "Description",
    "Vendor Name": "Vendor",
    "Location": "Location",
    "On Hand": "Quantity",
    "Reorder": "ReorderLevel",
    "Delete": "Discontinued"
    }   
 
    return value_mapping[input_value]

def fetch_inventory_data(description_text=None):
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()

    if description_text:
        cursor.execute("""
            SELECT 
                Inventory.ItemID, 
                Inventory.Description, 
                Vendors.VendorName, 
                Location.Location, 
                Location.SubLocation,
                Inventory.Quantity, 
                Inventory.ReorderLevel, 
                Inventory.Cost, 
                Inventory.Sell, 
                Inventory.Discontinued 
            FROM 
                Inventory 
            JOIN 
                Vendors ON Inventory.VendorID = Vendors.VendorID 
            JOIN 
                Location ON Inventory.LocationID = Location.LocationID 
            WHERE
                Inventory.Description LIKE ?
            ORDER BY 
                Inventory.Description, Vendors.VendorName;   
        """, ('%' + description_text + '%',))
    else:
        cursor.execute("""
            SELECT 
                Inventory.ItemID, 
                Inventory.Description, 
                Vendors.VendorName, 
                Location.Location, 
                Location.SubLocation,
                Inventory.Quantity, 
                Inventory.ReorderLevel, 
                Inventory.Cost, 
                Inventory.Sell, 
                Inventory.Discontinued 
            FROM 
                Inventory 
            JOIN 
                Vendors ON Inventory.VendorID = Vendors.VendorID 
            JOIN 
                Location ON Inventory.LocationID = Location.LocationID 
            ORDER BY 
                Inventory.Description, Vendors.VendorName;   
        """)

    data = cursor.fetchall()
    conn.close()
    return data

def read_location_data():
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    cursor.execute("SELECT LocationID, Location, SubLocation, Discontinued FROM Location ORDER BY Location")
    data = cursor.fetchall()
    conn.close()
    return data

def populate_treeview():
    items.delete(*items.get_children())  # Clear existing data
    data = fetch_inventory_data(search_text.get())
    for row in data:
        cost = "${:,.2f}".format(row[7] / 100)  # Index 6 corresponds to Cost
        sell = "${:,.2f}".format(row[8] / 100)  # Index 7 corresponds to Sell
        row_list = list(row)
        row_list[7] = cost
        row_list[8] = sell
        items.insert('', 'end', values=row_list, tags=("visible",))  # Initially mark all items as visible

def search_treeview():
    populate_treeview()
    sort_text.set('')
    sort_text.current(0)
    
def reset_treeview():
    search_text.delete(0,"end")
    populate_treeview()
    sort_text.set('')
    sort_text.current(0)

def sort_treeview_up(treeview, column_name):
    # Map column name to column index
    column_index = treeview['columns'].index(column_name)
    data = [(treeview.item(item)['values'][column_index], item) for item in treeview.get_children()]
    data.sort()  # Sort by specified column
    for index, (value, item) in enumerate(data):
        treeview.move(item, '', index)

def sort_treeview_dn(treeview, column_name):
    # Map column name to column index
    column_index = treeview['columns'].index(column_name)
    data = [(treeview.item(item)['values'][column_index], item) for item in treeview.get_children()]
    data.sort(reverse=True)  # Sort by specified column in reverse order
    for index, (value, item) in enumerate(data):
        treeview.move(item, '', index)

def delete_discontinued_zero_quantity():
    conn = sqlite3.connect("inventory.db")
    c = conn.cursor()
    sql_delete = """
    DELETE FROM Inventory
    WHERE Discontinued = 'Y' AND Quantity <= 0;
    """
    c.execute(sql_delete)
    conn.commit()
    conn.close()

def delete_discontinued_without_inventory():
    conn = sqlite3.connect("inventory.db")
    c = conn.cursor()
    sql_delete = """
    DELETE FROM Location
    WHERE Discontinued = 'Y' AND LocationID NOT IN (
        SELECT LocationID FROM Inventory
    );
    """
    c.execute(sql_delete)
    conn.commit()
    conn.close()

def delete_vendors_without_inventory():
    conn = sqlite3.connect("inventory.db")
    c = conn.cursor()
    sql_delete = """
    DELETE FROM Vendors
    WHERE Discontinued = 'Y' AND VendorID NOT IN (
        SELECT VendorID FROM Inventory
    );
    """
    c.execute(sql_delete)
    conn.commit()
    conn.close() 

def open_shortage_window():

    def read_shorts_data():
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                Inventory.Description, 
                Vendors.VendorName, 
                Vendors.RepName,
                Vendors.RepPhone,              
                Inventory.Quantity
            FROM 
                Inventory 
            JOIN 
                Vendors ON Inventory.VendorID = Vendors.VendorID 
            WHERE 
                Inventory.Quantity <= Inventory.ReorderLevel 
                AND 
                Inventory.Discontinued = 'N'
            ORDER BY 
                Vendors.VendorName, Inventory.Description;   
        """)
        data = cursor.fetchall()
        conn.close()
        return data
  
    shorts_window = tk.Toplevel(root)
    shorts_window.title("Inventory Shortage Report")
    center_window(shorts_window, 1000, 400)
    shorts_window.attributes("-topmost", True)  # Keep the window on top
    shorts_window.grab_set() 

    shorts_window.columnconfigure(0,weight=1)
    shorts_window.rowconfigure(0,weight=1)
    shorts_window.rowconfigure(1,weight=0)

    shorts_frame = tk.Frame(shorts_window)

    # Create Treeview widget
    shorts_treeview = ttk.Treeview(shorts_frame, columns=("Description", "VendorName", "RepName", "RepPhone","Quantity"), show="headings")
    shorts_treeview.heading("Description", text="Product Description")
    shorts_treeview.heading("VendorName", text="Vendor Name")
    shorts_treeview.heading("RepName", text="Rep Name")
    shorts_treeview.heading("RepPhone", text="Rep Phone")
    shorts_treeview.heading("Quantity", text="On Hand")

    # Hide VendorID column
    shorts_treeview["displaycolumns"] = ("Description", "VendorName", "RepName", "RepPhone", "Quantity")

    # Configure individual column settings
    # Configure individual column settings with minimum width
    shorts_treeview.column("Description", width=250, minwidth=250, anchor="center")  # Vendor Name width, centered
    shorts_treeview.column("VendorName", width=250, minwidth=250, anchor="center")  # Vendor Name width, centered
    shorts_treeview.column("RepName", width=100, minwidth=100, anchor="center")  # Rep Name width, centered
    shorts_treeview.column("RepPhone", width=100, minwidth=100, anchor="center")  # Rep Phone width, centered
    shorts_treeview.column("Quantity", width=20, minwidth=20, anchor="center")  # Delete width, centered

    # Create vertical scrollbar
    scrollbar = ttk.Scrollbar(shorts_frame, orient="vertical", command=shorts_treeview.yview)
     # Configure Treeview to use scrollbar
    shorts_treeview.configure(yscrollcommand=scrollbar.set)

    # Populate Treeview with vendor data
    shorts_data = read_shorts_data()
    for short in shorts_data:
        shorts_treeview.insert("", "end", values=short)

    shorts_cancel_button=ttk.Button(shorts_window, text="Close", width=15, command=lambda: shorts_window.destroy())
 
    # Pack
    shorts_treeview.pack(side="left", fill="both", expand=True)    
    scrollbar.pack(side="right", fill="y")
    shorts_frame.grid(row=0, column=0, sticky="nesw")
    shorts_cancel_button.grid(row=1, column=0, padx = 5, pady=10)
    shorts_cancel_button.focus()

def open_vendor_window():

    def read_vendors_data():
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("SELECT VendorID, VendorName, RepName, RepPhone, Discontinued FROM Vendors ORDER BY VendorName")
        data = cursor.fetchall()
        conn.close()
        return data

    def populate_vendor_fields(event):
        # Clear previous errors
        vendor_name_text.config(foreground="")
        vendor_rep_text.config(foreground="")
        vendor_ph_text.config(foreground="")

        # Get selected item
        selected_item = vendors_treeview.selection()
        if selected_item:
            # Retrieve values from selected item
            values = vendors_treeview.item(selected_item)['values']
            # Populate entry fields
            vendor_name_entry.delete(0, 'end')
            vendor_name_entry.insert(0, values[1])  
            vendor_rep_entry.delete(0, 'end')
            vendor_rep_entry.insert(0, values[2])   
            vendor_ph_entry.delete(0, 'end')
            vendor_ph_entry.insert(0, values[3])    
            if values[4] == "Y":
                vendor_del_var.set("Y")
            else:
                vendor_del_var.set("N")

    def clear_vendor_fields():
        # Clear entry fields
        vendor_name_entry.delete(0, 'end')
        vendor_rep_entry.delete(0, 'end')
        vendor_ph_entry.delete(0, 'end')
        vendor_del_var.set("N")
        # Unselect row in treeview
        vendors_treeview.selection_remove(vendors_treeview.selection())
        #clear errors
        vendor_name_text.config(foreground="")
        vendor_rep_text.config(foreground="")
        vendor_ph_text.config(foreground="")

    def validate_vendor_add():
 
        # Clear previous errors
        vendor_name_text.config(foreground="")
        vendor_rep_text.config(foreground="")
        vendor_ph_text.config(foreground="")

        # Check if fields are blank or contain only spaces
        valid = True
        if not vendor_name_entry.get().strip():
            vendor_name_text.config(foreground="red")
            valid = False
        if not vendor_rep_entry.get().strip():
            vendor_rep_text.config(foreground="red")
            valid = False
        if not vendor_ph_entry.get().strip():
            vendor_ph_text.config(foreground="red")
            valid = False
        if valid:
            confirmation = messagebox.askyesno("Confirmation Add Vendor", "Are you sure you want to add this vendor?", parent=vendor_window)
            if confirmation:
                add_vendor_record()
        else:
            messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=vendor_window)

    def add_vendor_record():
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Vendors (VendorName, RepName, RepPhone, Discontinued) VALUES (?, ?, ?, ?)",
                       (vendor_name_entry.get().strip(), vendor_rep_entry.get().strip(),
                        vendor_ph_entry.get().strip(), vendor_del_var.get()))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Vendor record added successfully.", parent=vendor_window)

        # Refresh vendor data in Treeview
        refresh_vendor_data()

    def validate_vendor_modify():
 
        selected_item = vendors_treeview.selection()
        if selected_item:
            # Clear previous errors
            vendor_name_entry.config(foreground="")
            vendor_rep_entry.config(foreground="")
            vendor_ph_entry.config(foreground="")

            # Check if fields are blank or contain only spaces
            valid = True
            if not vendor_name_entry.get().strip():
                vendor_name_text.config(foreground="red")
                valid = False
            if not vendor_rep_entry.get().strip():
                vendor_rep_text.config(foreground="red")
                valid = False
            if not vendor_ph_entry.get().strip():
                vendor_ph_text.config(foreground="red")
                valid = False
            if valid:
                confirmation = messagebox.askyesno("Confirmation Vendor Changes", "Are you sure you want to save changes to this vendor?", parent=vendor_window)
                if confirmation:
                    modify_vendor_record()
            else:
                messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=vendor_window)
        else:
            messagebox.showwarning("No Vendor Selected", "Please select a vendor to modify.", parent=vendor_window)

    def update_vendor_data(vendor_id, vendor_name, rep_name, rep_phone, discontinued):
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("UPDATE Vendors SET VendorName=?, RepName=?, RepPhone=?, Discontinued=? WHERE VendorID=?", (vendor_name, rep_name, rep_phone, discontinued, vendor_id))
        conn.commit()
        conn.close()

    def modify_vendor_record():
        selected_item = vendors_treeview.selection()
        if selected_item:
            # Get selected item
            # Retrieve values from selected item
            values = vendors_treeview.item(selected_item)['values']
            vendor_id = values[0]
            vendor_name = vendor_name_entry.get().strip()
            rep_name = vendor_rep_entry.get().strip()
            rep_phone = vendor_ph_entry.get().strip()
            discontinued = vendor_del_var.get()
            # Update vendor data
            update_vendor_data(vendor_id, vendor_name, rep_name, rep_phone, discontinued)
            messagebox.showinfo("Success", "Vendor updated successfully.", parent = vendor_window)
            # Refresh vendor data in Treeview
            refresh_vendor_data()
 
    def refresh_vendor_data():
        # Clear entry fields and errors
        clear_vendor_fields()
        # Reload vendor data into Treeview
        vendors_treeview.delete(*vendors_treeview.get_children())
        vendors_data = read_vendors_data()
        for vendor in vendors_data:
            vendors_treeview.insert("", "end", values=vendor)
   
    def close_vendor_window():
        reset_treeview()
        vendor_window.destroy()

    vendor_window = tk.Toplevel(root)
    vendor_window.title("Modify Product Vendors")
    center_window(vendor_window, 800, 400)    
    vendor_window.attributes("-topmost", True)  # Keep the window on top
    vendor_window.grab_set() 

    vendor_window.grid_columnconfigure((0, 1, 2, 3), weight=1, uniform="equal")
    vendor_window.rowconfigure(0,weight=1)
    vendor_window.rowconfigure(1,weight=0)
    vendor_window.rowconfigure(2,weight=0)
    vendor_window.rowconfigure(3,weight=0)

    vendors_frame = tk.Frame(vendor_window)

    # Create Treeview widget
    vendors_treeview = ttk.Treeview(vendors_frame, columns=("VendorID", "VendorName", "RepName", "RepPhone","Discontinued"), show="headings")
    vendors_treeview.heading("#0", text="VendorID")
    vendors_treeview.heading("VendorName", text="Vendor Name")
    vendors_treeview.heading("RepName", text="Rep Name")
    vendors_treeview.heading("RepPhone", text="Rep Phone")
    vendors_treeview.heading("Discontinued", text="Delete")

    # Hide VendorID column
    vendors_treeview["displaycolumns"] = ("VendorName", "RepName", "RepPhone", "Discontinued")
    vendors_treeview.column("#0", width=0, stretch=False)

    # Configure individual column settings
    # Configure individual column settings with minimum width

    vendors_treeview.column("VendorName", width=250, minwidth=250, anchor="center")  # Vendor Name width, centered
    vendors_treeview.column("RepName", width=110, minwidth=110, anchor="center")  # Rep Name width, centered
    vendors_treeview.column("RepPhone", width=100, minwidth=100, anchor="center")  # Rep Phone width, centered
    vendors_treeview.column("Discontinued", width=20, minwidth=20, anchor="center")  # Delete width, centered

    # Create vertical scrollbar
    scrollbar = ttk.Scrollbar(vendors_frame, orient="vertical", command=vendors_treeview.yview)
 
    # Configure Treeview to use scrollbar
    vendors_treeview.configure(yscrollcommand=scrollbar.set)

    # Populate Treeview with vendor data
    vendors_data = read_vendors_data()
    for vendor in vendors_data:
        vendors_treeview.insert("", "end", values=vendor)

    vendor_name_text=ttk.Label(vendor_window, text = "Vendor Name")
    vendor_name_entry=ttk.Entry(vendor_window, width = 40)
    vendor_del_text=ttk.Label(vendor_window, text="Delete")
    vendor_del_var = tk.StringVar(value="N")
    vendor_del_entry = ttk.Checkbutton(vendor_window, variable=vendor_del_var, onvalue="Y", offvalue="N")    
    vendor_rep_text=ttk.Label(vendor_window, text="Rep Name")
    vendor_rep_entry=ttk.Entry(vendor_window, width = 40)
    vendor_ph_text=ttk.Label(vendor_window, text="Rep Phone")
    vendor_ph_entry=ttk.Entry(vendor_window, width=20)
    vendor_clear_button=ttk.Button(vendor_window, text="Clear", width=15, command=clear_vendor_fields)
    vendor_add_button=ttk.Button(vendor_window, text="Add Vendor", width = 15, command=validate_vendor_add)
    vendor_modify_button=ttk.Button(vendor_window, text="Apply Changes", width = 15, command=validate_vendor_modify)
    vendor_cancel_button=ttk.Button(vendor_window, text="Close", width=15, command=close_vendor_window)
 

    vendors_treeview.bind('<ButtonRelease-1>', populate_vendor_fields)

    # Pack
    vendors_treeview.pack(side="left", fill="both", expand=True)    
    scrollbar.pack(side="right", fill="y")
    vendors_frame.grid(row=0, column=0, columnspan=4, sticky="nesw")
    vendor_name_text.grid(row=1, column=0, padx = 5, sticky="e")
    vendor_name_entry.grid(row=1, column=1, sticky="w")
    vendor_del_text.grid(row=1, column=2, padx = 5, sticky="e")
    vendor_del_entry.grid(row=1, column=3, sticky="w")
    vendor_rep_text.grid(row=2, column=0, padx = 5, sticky="e")
    vendor_rep_entry.grid(row=2, column=1, sticky="w")
    vendor_ph_text.grid(row=2, column=2, padx = 5, sticky="e")
    vendor_ph_entry.grid(row=2, column=3, sticky="w")
    vendor_cancel_button.grid(row=3, column=3, padx = 5, pady=10)
    vendor_modify_button.grid(row=3, column=2, padx = 5, pady=10)
    vendor_add_button.grid(row=3, column=1, padx = 5, pady=10)
    vendor_clear_button.grid(row=3, column=0, padx=5, pady=10)

    vendor_name_entry.focus()

def open_locations_window():

    def read_locations_data():
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("SELECT LocationID, Location, SubLocation, Discontinued FROM Location ORDER BY Location")
        data = cursor.fetchall()
        conn.close()
        return data

    def populate_location_fields(event):
        # Clear previous errors
        location_name_text.config(foreground="")

        # Get selected item
        selected_item = locations_treeview.selection()
        if selected_item:
            # Retrieve values from selected item
            values = locations_treeview.item(selected_item)['values']
            # Populate entry fields
            location_name_entry.delete(0, 'end')
            location_name_entry.insert(0, values[1])  
            location_sub_entry.delete(0, 'end')
            location_sub_entry.insert(0, values[2])   
            if values[3] == "Y":
                location_del_var.set("Y")
            else:
                location_del_var.set("N")

    def clear_location_fields():
        # Clear entry fields
        location_name_entry.delete(0, 'end')
        location_sub_entry.delete(0, 'end')
        location_del_var.set("N")
        locations_treeview.selection_remove(locations_treeview.selection())
        location_name_text.config(foreground="")
        location_sub_text.config(foreground="")
 
    def validate_location_add():
 
        # Clear previous errors
        location_name_text.config(foreground="")
        location_sub_text.config(foreground="")
 
        # Check if fields are blank or contain only spaces
        valid = True
        if not location_name_entry.get().strip():
            location_name_text.config(foreground="red")
            valid = False
        if not location_sub_entry.get().strip():
            location_sub_text.config(foreground="red")
            valid = False
        if valid:
            confirmation = messagebox.askyesno("Confirmation Add Location", "Are you sure you want to add this location?", parent=location_window)
            if confirmation:
                add_location_record()
        else:
            messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=location_window)

    def add_location_record():
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Location (Location, SubLocation, Discontinued) VALUES (?, ?, ?)",
                       (location_name_entry.get().strip(), location_sub_entry.get().strip(), location_del_var.get()))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Location added successfully.", parent=location_window)

        # Refresh location data in Treeview
        refresh_location_data()

    def validate_location_modify():
 
        selected_item = locations_treeview.selection()
        if selected_item:
            # Clear previous errors
            location_name_text.config(foreground="")
            location_sub_text.config(foreground="")

            # Check if fields are blank or contain only spaces
            valid = True
            if not location_name_entry.get().strip():
                location_name_text.config(foreground="red")
                valid = False
            if not location_sub_entry.get().strip():
                location_sub_text.config(foreground="red")
                valid = False

            if valid:
                confirmation = messagebox.askyesno("Confirm Location Changes", "Are you sure you want to save changes to this location?", parent=location_window)
                if confirmation:
                    modify_location_record()
            else:
                messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=location_window)
        else:
            messagebox.showwarning("No Location Selected", "Please select a location to modify.", parent=location_window)

    def update_location_data(location_id, location_name, sublocation, discontinued):
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("UPDATE Location SET Location=?, SubLocation=?, Discontinued=? WHERE LocationID=?", (location_name, sublocation, discontinued, location_id))
        conn.commit()
        conn.close()

    def modify_location_record():
        selected_item = locations_treeview.selection()
        if selected_item:
            # Get selected item
            # Retrieve values from selected item
            values = locations_treeview.item(selected_item)['values']
            location_id = values[0]
            location_name = location_name_entry.get().strip()
            sublocation = location_sub_entry.get().strip()
            discontinued = location_del_var.get()
            # Update location data
            update_location_data(location_id, location_name, sublocation, discontinued)
            messagebox.showinfo("Success", "Location updated successfully.", parent = location_window)
            # Refresh location data in Treeview
            refresh_location_data()
 
    def refresh_location_data():
        # Clear entry fields and errors
        clear_location_fields()
        # Reload location data into Treeview
        locations_treeview.delete(*locations_treeview.get_children())
        locations_data = read_locations_data()
        for location in locations_data:
            locations_treeview.insert("", "end", values=location)
   
    def close_location_window():
        reset_treeview()
        location_window.destroy()

    location_window = tk.Toplevel(root)
    location_window.title("Modify Product Locations")
    center_window(location_window, 800, 400)    
    location_window.attributes("-topmost", True)  # Keep the window on top
    location_window.grab_set()   

    location_window.grid_columnconfigure((0, 1, 2, 3), weight=1, uniform="equal")
    location_window.rowconfigure(0,weight=1)
    location_window.rowconfigure(1,weight=0)
    location_window.rowconfigure(2,weight=0)
    location_window.rowconfigure(3,weight=0)

    locations_frame = tk.Frame(location_window)

    # Create Treeview widget
    locations_treeview = ttk.Treeview(locations_frame, columns=("LocationID", "Location", "SubLocation", "Discontinued"), show="headings")
    locations_treeview.heading("#0", text="LocationID")
    locations_treeview.heading("Location", text="Location Name")
    locations_treeview.heading("SubLocation", text="Sub Location")
    locations_treeview.heading("Discontinued", text="Delete")

    # Hide locationID column
    locations_treeview["displaycolumns"] = ("Location", "SubLocation", "Discontinued")
    locations_treeview.column("#0", width=0, stretch=False)

    # Configure individual column settings
    # Configure individual column settings with minimum width

    locations_treeview.column("Location", width=140, minwidth=140, anchor="center")  # location Name width, centered
    locations_treeview.column("SubLocation", width=125, minwidth=125, anchor="center")  # Rep Name width, centered
    locations_treeview.column("Discontinued", width=40, minwidth=40, anchor="center")  # Delete width, centered

    # Create vertical scrollbar
    scrollbar = ttk.Scrollbar(locations_frame, orient="vertical", command=locations_treeview.yview)
 

    # Configure Treeview to use scrollbar
    locations_treeview.configure(yscrollcommand=scrollbar.set)

    # Populate Treeview with location data
    locations_data = read_locations_data()
    for location in locations_data:
        locations_treeview.insert("", "end", values=location)

    location_name_text=ttk.Label(location_window, text = "Location")
    location_name_entry=ttk.Entry(location_window, width = 30)
    location_del_text=ttk.Label(location_window, text="Delete")
    location_del_var = tk.StringVar(value="N")
    location_del_entry = ttk.Checkbutton(location_window, variable=location_del_var, onvalue="Y", offvalue="N")    
    location_sub_text=ttk.Label(location_window, text="Sub-Location")
    location_sub_entry=ttk.Entry(location_window, width = 30)
    location_clear_button=ttk.Button(location_window, text="Clear", width=15, command=clear_location_fields)
    location_add_button=ttk.Button(location_window, text="Add Location", width = 15, command=validate_location_add)
    location_modify_button=ttk.Button(location_window, text="Apply Changes", width = 15, command=validate_location_modify)
    location_cancel_button=ttk.Button(location_window, text="Close", width=15, command=close_location_window)
 

    locations_treeview.bind('<ButtonRelease-1>', populate_location_fields)

    # Pack
    locations_treeview.pack(side="left", fill="both", expand=True)    
    scrollbar.pack(side="right", fill="y")
    locations_frame.grid(row=0, column=0, columnspan=4, sticky="nesw")
    location_name_text.grid(row=1, column=0, padx = 5, sticky="e")
    location_name_entry.grid(row=1, column=1, sticky="w")
    location_del_text.grid(row=1, column=2, padx = 5, sticky="e")
    location_del_entry.grid(row=1, column=3, sticky="w")
    location_sub_text.grid(row=2, column=0, padx = 5, sticky="e")
    location_sub_entry.grid(row=2, column=1, sticky="w")
    location_cancel_button.grid(row=3, column=3, padx = 5, pady=10)
    location_modify_button.grid(row=3, column=2, padx = 5, pady=10)
    location_add_button.grid(row=3, column=1, padx = 5, pady=10)
    location_clear_button.grid(row=3, column=0, padx=5, pady=10)
    location_name_entry.focus()

def add_item_window():
    add_item_window = tk.Toplevel(root)
    add_item_window.title("Add New Product")
    center_window(add_item_window, 700, 400)    
    add_item_window.attributes("-topmost", True)  # Keep the window on top
    add_item_window.grab_set() 

    add_item_window.grid_columnconfigure((0, 1, 2), weight=1, uniform="equal")
    add_item_window.grid_rowconfigure((0,1,2,3,4,5,6,7), weight=1)
    add_item_window.rowconfigure(8,weight=0)

    def validate_input():
        # Validation for each input field
        fail = False
        add_item_description_label.configure(foreground="")
        add_item_location_label.configure(foreground="")
        add_item_vendor_label.configure(foreground="")
        add_item_cost_label.configure(foreground="")
        add_item_sell_label.configure(foreground="")
        add_item_reorder_label.configure(foreground="")
        add_item_quantity_label.configure(foreground="")

        if not add_item_description_entry.get():
            add_item_description_label.config(foreground="red")       
            fail = True
        if not add_item_vendor_combobox.get():
            add_item_vendor_label.config(foreground="red")
            fail = True
        if not add_item_location_combobox.get():
            add_item_location_label.config(foreground="red")
            fail = True
        try:
            float(add_item_cost_entry.get().replace(',', '').replace('$', '').strip())
        except ValueError:
            add_item_cost_label.config(foreground="red")
            fail = True
        try:
            float(add_item_sell_entry.get().replace(',', '').replace('$', '').strip())
        except ValueError:
           add_item_sell_label.config(foreground="red")
           fail = True
        try:
            int(add_item_reorder_entry.get())
        except ValueError:
            add_item_reorder_label.config(foreground="red")
            fail =True
        try:
            int(add_item_quantity_entry.get())
        except ValueError:
            add_item_quantity_label.config(foreground="red")
            fail = True
        if fail :
            messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=add_item_window)
            return False
        confirmation = messagebox.askyesno("Confirm Product Addition", "Are you sure you want to add this product?", parent=add_item_window)
        if confirmation:
            return True
        return False

    def populate_comboboxes():
        # Connect to the database
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        # Retrieve vendors and locations from the database
        c.execute("SELECT VendorName FROM Vendors WHERE Discontinued = 'N'")
        vendors = [row[0] for row in c.fetchall()]
        add_item_vendor_combobox["values"] = vendors

        c.execute("SELECT Location, SubLocation FROM Location WHERE Discontinued = 'N'")
        locations = [f"{row[0]} - {row[1]}" if row[1] else row[0] for row in c.fetchall()]
        add_item_location_combobox["values"] = locations

        # Close the connection
        conn.close()

    def add_item():
        if not validate_input():
            return
        # Retrieve values from the entry fields and comboboxes
        description = add_item_description_entry.get()
        vendor = add_item_vendor_combobox.get()
        location = add_item_location_combobox.get().split(" - ")[0]  # Split to get only location
        cost = add_item_cost_entry.get().replace(',', '').replace('$', '').strip()
        sell = add_item_sell_entry.get().replace(',', '').replace('$', '').strip()       
        reorder_level = add_item_reorder_entry.get().replace(',', '').strip()  
        discontinued = add_item_discontinued_var.get()
        quantity = add_item_quantity_entry.get().replace(',', '').strip()

        # Insert the new item into the database
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        c.execute("INSERT INTO Inventory (Description, VendorID, LocationID, Quantity, Cost, Sell, ReorderLevel, Discontinued) VALUES (?, (SELECT VendorID FROM Vendors WHERE VendorName = ?), (SELECT LocationID FROM Location WHERE Location = ?), ?, ?, ?, ?, ?)",
                (description, vendor, location, int(quantity), int(float(cost) * 100), int(float(sell) * 100), int(reorder_level), discontinued))

        conn.commit()
        conn.close()
        clear_item_entry()
        messagebox.showinfo("Success", "Product added successfully.", parent=add_item_window)

    def clear_item_entry():
        # Clear entry fields after adding item
        add_item_description_label.configure(foreground="")
        add_item_location_label.configure(foreground="")
        add_item_vendor_label.configure(foreground="")
        add_item_cost_label.configure(foreground="")
        add_item_sell_label.configure(foreground="")
        add_item_reorder_label.configure(foreground="")
        add_item_quantity_label.configure(foreground="")

        add_item_description_entry.delete(0, tk.END)
        add_item_vendor_combobox.set('')
        add_item_location_combobox.set('')
        add_item_cost_entry.delete(0, tk.END)
        add_item_cost_entry.insert(0, "0.00")
        add_item_sell_entry.delete(0, tk.END)
        add_item_sell_entry.insert(0, "0.00")
        add_item_reorder_entry.delete(0, tk.END)
        add_item_reorder_entry.insert(0, "0")
        add_item_discontinued_var.set('N')
        add_item_quantity_entry.delete(0, "0")
        add_item_description_entry.focus()
       
    def close_add_item_window():
        reset_treeview()
        add_item_window.destroy()

    add_item_description_label = ttk.Label(add_item_window, text="Product Description ")
    add_item_description_label.grid(row=0, column=0, sticky="e")
    add_item_description_entry = ttk.Entry(add_item_window, width = 60)
    add_item_description_entry.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    add_item_vendor_label = ttk.Label(add_item_window, text="Vendor Name ")
    add_item_vendor_label.grid(row=1, column=0, sticky="e")
    add_item_vendor_combobox = ttk.Combobox(add_item_window, width = 40, state="readonly")
    add_item_vendor_combobox.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    add_item_location_label = ttk.Label(add_item_window, text="Location ")
    add_item_location_label.grid(row=2, column=0, sticky="e")
    add_item_location_combobox = ttk.Combobox(add_item_window, state="readonly", width=50)
    add_item_location_combobox.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    add_item_cost_label = ttk.Label(add_item_window, text="Cost Price ")
    add_item_cost_label.grid(row=3, column=0, sticky="e")
    add_item_cost_entry = ttk.Entry(add_item_window, width = 20)
    add_item_cost_entry.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    add_item_cost_entry.insert(0, "0.00")

    add_item_sell_label = ttk.Label(add_item_window, text="Sell Price ")
    add_item_sell_label.grid(row=4, column=0, sticky="e")
    add_item_sell_entry = ttk.Entry(add_item_window, width=20)
    add_item_sell_entry.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    add_item_sell_entry.insert(0, "0.00")

    add_item_reorder_label = ttk.Label(add_item_window, text="Re-Order Level ")
    add_item_reorder_label.grid(row=5, column=0, sticky="e")
    add_item_reorder_entry = ttk.Entry(add_item_window)
    add_item_reorder_entry.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    add_item_reorder_entry.insert(0, "0")
 
    add_item_quantity_label = ttk.Label(add_item_window, text="On Hand ")
    add_item_quantity_label.grid(row=6, column=0, sticky="e")
    add_item_quantity_entry = ttk.Entry(add_item_window, width=20)
    add_item_quantity_entry.grid(row=6, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    add_item_quantity_entry.insert(0, "0")

    add_item_discontinued_label = ttk.Label(add_item_window, text="Discontinued ")
    add_item_discontinued_label.grid(row=7, column=0, sticky="e")
    add_item_discontinued_var = tk.StringVar(value='N')
    add_item_discontinued_checkbutton = ttk.Checkbutton(add_item_window, variable=add_item_discontinued_var, onvalue="Y", offvalue="N")     
    add_item_discontinued_checkbutton.grid(row=7, column=1, columnspan=2, pady=5, sticky="w")
     
    add_item_clear_button=ttk.Button(add_item_window, text="Clear", width=15, command=clear_item_entry)
    add_item_add_button=ttk.Button(add_item_window, text="Add Product", width = 15, command=add_item)
    add_item_cancel_button=ttk.Button(add_item_window, text="Close", width=15, command=close_add_item_window)
 
    add_item_cancel_button.grid(row=8, column=2, padx = 5, pady=10)
    add_item_add_button.grid(row=8, column=1, padx = 5, pady=10)
    add_item_clear_button.grid(row=8, column=0, padx=5, pady=10)
    add_item_description_entry.focus()

    populate_comboboxes()

def open_edit_item_window():
    selected_item = items.selection()
    if selected_item:
        edit_item_window(items.item(items.selection()[0], 'values')[0])
    else:
        messagebox.showwarning("No Product Selected", "Please select a product to modify.", parent=root)       

def transaction_window(ItemID):
    transaction_window = tk.Toplevel(root)
    transaction_window.title("Enter Product Transaction")
    center_window(transaction_window, 500, 200)    
    transaction_window.attributes("-topmost", True)  # Keep the window on top
    transaction_window.grab_set() 

    transaction_window.grid_columnconfigure((0, 1), weight=1, uniform="equal")
    transaction_window.grid_rowconfigure((0,1,2,3,4,5,6), weight=1)
    transaction_window.rowconfigure(7,weight=0)

    def fetch_item_data(ItemID):
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        # Fetch item data based on ItemID
        c.execute("""SELECT Inventory.Description, Vendors.VendorName, Location.Location, 
                    Location.SubLocation, Inventory.Quantity
             FROM Inventory 
             INNER JOIN Vendors ON Inventory.VendorID = Vendors.VendorID 
             INNER JOIN Location ON Inventory.LocationID = Location.LocationID
             WHERE Inventory.ItemID = ?""", (ItemID,))
        item_data = c.fetchone()

        conn.close()
        return item_data

    def populate_entry_fields(item_data):
        if item_data:
            transaction_description_label.configure(text=item_data[0])
            transaction_vendor_label.configure(text=item_data[1])
            transaction_location_label.configure(text=item_data[2] + " - " + item_data[3])
            transaction_quantity_label.configure(text="On-Hand: " + str(item_data[4]))
            return item_data[4]

    def validate_input():
        # Validation for each input field
        fail = False
        transaction_purchase_label.configure(foreground="")
        transaction_sell_label.configure(foreground="")

        try:
            int(transaction_purchase_entry.get())
        except ValueError:
            transaction_purchase_label.config(foreground="red")
            fail = True
        try:
            int(transaction_sell_entry.get())
        except ValueError:
           transaction_sell_label.config(foreground="red")
           fail = True
        if fail :
            messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=transaction_window)
            return False
        confirmation = messagebox.askyesno("Confirm Quantity Changes", "Are you sure you want to change the quantity?", parent=transaction_window)
        if confirmation:
            return True
        return False

    def edit_item():
        if not validate_input():
            return
        quantity = on_hand_quantity + int(transaction_purchase_entry.get()) - int(transaction_sell_entry.get())

        # Insert the new item into the database
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        c.execute("""UPDATE Inventory 
                    SET Quantity = ? 
                    WHERE ItemID = ?""",
                (int(quantity), ItemID))

        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Quantity updated successfully.", parent=transaction_window)
        transaction_window.destroy()
        reset_treeview()

       
    transaction_description_label = ttk.Label(transaction_window, text="Product Description")
    transaction_description_label.grid(row=0, column=0, columnspan=2)

    transaction_vendor_label = ttk.Label(transaction_window, text="Vendor Name")
    transaction_vendor_label.grid(row=1, column=0, columnspan=2)

    transaction_location_label = ttk.Label(transaction_window, text="Location")
    transaction_location_label.grid(row=2, column=0, columnspan=2)
 
    transaction_quantity_label = ttk.Label(transaction_window, text="On Hand")
    transaction_quantity_label.grid(row=3, column=0, columnspan=2)

    transaction_space_label = ttk.Label(transaction_window, text=" ")
    transaction_space_label.grid(row=4, column=0, columnspan=2)

    transaction_purchase_label = ttk.Label(transaction_window, text="Purchased")
    transaction_purchase_label.grid(row=5, column=0)
    transaction_purchase_entry = ttk.Entry(transaction_window, width=10)
    transaction_purchase_entry.grid(row=6, column=0, padx=5, pady=5)
    transaction_purchase_entry.insert(0, "0")

    transaction_sell_label = ttk.Label(transaction_window, text="Sold")
    transaction_sell_label.grid(row=5, column=1)
    transaction_sell_entry = ttk.Entry(transaction_window, width = 10)
    transaction_sell_entry.grid(row=6, column=1, padx=5, pady=5)
    transaction_sell_entry.insert(0, "0")
  
    transaction_edit_button=ttk.Button(transaction_window, text="Apply Changes", width = 15, command=edit_item)
    transaction_cancel_button=ttk.Button(transaction_window, text="Close", width=15, command=lambda: transaction_window.destroy())
 
    transaction_cancel_button.grid(row=7, column=1, padx = 5, pady=10)
    transaction_edit_button.grid(row=7, column=0, padx = 5, pady=10)
    transaction_purchase_entry.focus()

    item_data = fetch_item_data(ItemID)
    on_hand_quantity = int(populate_entry_fields(item_data))

def open_transaction_window():
    selected_item = items.selection()
    if selected_item:
        transaction_window(items.item(items.selection()[0], 'values')[0])
    else:
        messagebox.showwarning("No Product Selected", "Please select a product to modify.", parent=root) 

def edit_item_window(ItemID):
    edit_item_window = tk.Toplevel(root)
    edit_item_window.title("Edit Existing Product")
    center_window(edit_item_window, 700, 400)   
    edit_item_window.attributes("-topmost", True)  # Keep the window on top
    edit_item_window.grab_set() 

    edit_item_window.grid_columnconfigure((0, 1, 2), weight=1, uniform="equal")
    edit_item_window.grid_rowconfigure((0,1,2,3,4,5,6,7), weight=1)
    edit_item_window.rowconfigure(8,weight=0)

    def fetch_item_data(ItemID):
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        # Fetch item data based on ItemID
        c.execute("""SELECT Inventory.Description, Vendors.VendorName, Location.Location, 
                    Location.SubLocation, Inventory.Quantity, Inventory.Cost, Inventory.Sell, 
                    Inventory.ReorderLevel, Inventory.Discontinued
             FROM Inventory 
             INNER JOIN Vendors ON Inventory.VendorID = Vendors.VendorID 
             INNER JOIN Location ON Inventory.LocationID = Location.LocationID
             WHERE Inventory.ItemID = ?""", (ItemID,))
        item_data = c.fetchone()

        conn.close()
        return item_data

    def populate_entry_fields(item_data):
        if item_data:
            # Populate entry fields with fetched data
            edit_item_description_entry.delete(0, tk.END)
            edit_item_description_entry.insert(0, item_data[0])

            edit_item_vendor_combobox.set(item_data[1])

            edit_item_location_combobox.set(f"{item_data[2]} - {item_data[3]}")

            edit_item_cost_entry.delete(0, tk.END)
            edit_item_cost_entry.insert(0, f"${item_data[5] / 100:.2f}")

            edit_item_sell_entry.delete(0, tk.END)
            edit_item_sell_entry.insert(0, f"${item_data[6] / 100:.2f}")

            edit_item_reorder_entry.delete(0, tk.END)
            edit_item_reorder_entry.insert(0, item_data[7])

            edit_item_quantity_entry.delete(0, tk.END)
            edit_item_quantity_entry.insert(0, item_data[4])

            edit_item_discontinued_var.set(item_data[8])


    def validate_input():
        # Validation for each input field
        fail = False
        edit_item_description_label.configure(foreground="")
        edit_item_location_label.configure(foreground="")
        edit_item_vendor_label.configure(foreground="")
        edit_item_cost_label.configure(foreground="")
        edit_item_sell_label.configure(foreground="")
        edit_item_reorder_label.configure(foreground="")
        edit_item_quantity_label.configure(foreground="")

        if not edit_item_description_entry.get():
            edit_item_description_label.config(foreground="red")       
            fail = True
        if not edit_item_vendor_combobox.get():
            edit_item_vendor_label.config(foreground="red")
            fail = True
        if not edit_item_location_combobox.get():
            edit_item_location_label.config(foreground="red")
            fail = True
        try:
            float(edit_item_cost_entry.get().replace(',', '').replace('$', '').strip())
        except ValueError:
            edit_item_cost_label.config(foreground="red")
            fail = True
        try:
            float(edit_item_sell_entry.get().replace(',', '').replace('$', '').strip())
        except ValueError:
           edit_item_sell_label.config(foreground="red")
           fail = True
        try:
            int(edit_item_reorder_entry.get())
        except ValueError:
            edit_item_reorder_label.config(foreground="red")
            fail =True
        try:
            int(edit_item_quantity_entry.get())
        except ValueError:
            edit_item_quantity_label.config(foreground="red")
            fail = True
        if fail :
            messagebox.showwarning("Missing Required Information", "Please enter all required information.", parent=edit_item_window)
            return False
        confirmation = messagebox.askyesno("Confirm Product Changes", "Are you sure you want to modify this product?", parent=edit_item_window)
        if confirmation:
            return True
        return False

    def populate_comboboxes():
        # Connect to the database
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        # Retrieve vendors and locations from the database
        c.execute("SELECT VendorName FROM Vendors")
        vendors = [row[0] for row in c.fetchall()]
        edit_item_vendor_combobox["values"] = vendors

        c.execute("SELECT Location, SubLocation FROM Location")
        locations = [f"{row[0]} - {row[1]}" if row[1] else row[0] for row in c.fetchall()]
        edit_item_location_combobox["values"] = locations

        # Close the connection
        conn.close()

    def edit_item():
        if not validate_input():
            return
        # Retrieve values from the entry fields and comboboxes
        description = edit_item_description_entry.get()
        vendor = edit_item_vendor_combobox.get()
        location = edit_item_location_combobox.get().split(" - ")[0]  # Split to get only location
        cost = edit_item_cost_entry.get().replace(',', '').replace('$', '').strip()
        sell = edit_item_sell_entry.get().replace(',', '').replace('$', '').strip()       
        reorder_level = edit_item_reorder_entry.get().replace(',', '').strip()  
        discontinued = edit_item_discontinued_var.get()
        quantity = edit_item_quantity_entry.get().replace(',', '').strip()

        # Insert the new item into the database
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        c.execute("""UPDATE Inventory 
                        SET Description = ?, 
                            VendorID = (SELECT VendorID FROM Vendors WHERE VendorName = ?), 
                            LocationID = (SELECT LocationID FROM Location WHERE Location = ?), 
                            Quantity = ?, 
                            Cost = ?, 
                            Sell = ?, 
                            ReorderLevel = ?, 
                            Discontinued = ? 
                        WHERE ItemID = ?""",
                    (description, vendor, location, int(quantity), int(float(cost) * 100), int(float(sell) * 100), int(reorder_level), discontinued, ItemID))

        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Product updated successfully.", parent=edit_item_window)
        edit_item_window.destroy()
        reset_treeview()

       
    edit_item_description_label = ttk.Label(edit_item_window, text="Product Description ")
    edit_item_description_label.grid(row=0, column=0, sticky="e")
    edit_item_description_entry = ttk.Entry(edit_item_window, width = 60)
    edit_item_description_entry.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    edit_item_vendor_label = ttk.Label(edit_item_window, text="Vendor Name ")
    edit_item_vendor_label.grid(row=1, column=0, sticky="e")
    edit_item_vendor_combobox = ttk.Combobox(edit_item_window, width = 40, state="readonly")
    edit_item_vendor_combobox.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    edit_item_location_label = ttk.Label(edit_item_window, text="Location ")
    edit_item_location_label.grid(row=2, column=0, sticky="e")
    edit_item_location_combobox = ttk.Combobox(edit_item_window, state="readonly", width=50)
    edit_item_location_combobox.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    edit_item_cost_label = ttk.Label(edit_item_window, text="Cost Price ")
    edit_item_cost_label.grid(row=3, column=0, sticky="e")
    edit_item_cost_entry = ttk.Entry(edit_item_window, width = 20)
    edit_item_cost_entry.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    edit_item_cost_entry.insert(0, "0.00")

    edit_item_sell_label = ttk.Label(edit_item_window, text="Sell Price ")
    edit_item_sell_label.grid(row=4, column=0, sticky="e")
    edit_item_sell_entry = ttk.Entry(edit_item_window, width=20)
    edit_item_sell_entry.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    edit_item_sell_entry.insert(0, "0.00")

    edit_item_reorder_label = ttk.Label(edit_item_window, text="Re-Order Level ")
    edit_item_reorder_label.grid(row=5, column=0, sticky="e")
    edit_item_reorder_entry = ttk.Entry(edit_item_window)
    edit_item_reorder_entry.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    edit_item_reorder_entry.insert(0, "0")
 
    edit_item_quantity_label = ttk.Label(edit_item_window, text="On Hand ")
    edit_item_quantity_label.grid(row=6, column=0, sticky="e")
    edit_item_quantity_entry = ttk.Entry(edit_item_window, width=20)
    edit_item_quantity_entry.grid(row=6, column=1, columnspan=2, padx=5, pady=5, sticky="w")
    edit_item_quantity_entry.insert(0, "0")

    edit_item_discontinued_label = ttk.Label(edit_item_window, text="Discontinued ")
    edit_item_discontinued_label.grid(row=7, column=0, sticky="e")
    edit_item_discontinued_var = tk.StringVar(value='N')
    edit_item_discontinued_checkbutton = ttk.Checkbutton(edit_item_window, variable=edit_item_discontinued_var, onvalue="Y", offvalue="N")     
    edit_item_discontinued_checkbutton.grid(row=7, column=1, columnspan=2, pady=5, sticky="w")
     
    edit_item_edit_button=ttk.Button(edit_item_window, text="Apply Changes", width = 15, command=edit_item)
    edit_item_cancel_button=ttk.Button(edit_item_window, text="Close", width=15, command=lambda: edit_item_window.destroy())
 
    edit_item_cancel_button.grid(row=8, column=2, padx = 5, pady=10)
    edit_item_edit_button.grid(row=8, column=1, padx = 5, pady=10)
    edit_item_description_entry.focus()

    populate_comboboxes()
    item_data = fetch_item_data(ItemID)
    populate_entry_fields(item_data)

def edit_settings_window():
    edit_settings_window = tk.Toplevel(root)
    edit_settings_window.title("Edit Discounts")
    center_window(edit_settings_window, 400, 270)
    edit_settings_window.attributes("-topmost", True)  # Keep the window on top
    edit_settings_window.grab_set() 

    edit_settings_window.grid_columnconfigure((0, 1), weight=1, uniform="equal")
    edit_settings_window.grid_rowconfigure((0,1,2,3,4,5), weight=1)
    edit_settings_window.rowconfigure(6,weight=0)

    def fetch_settings_data():
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()
        c.execute("SELECT * FROM Settings WHERE Purpose = 'Discounts'")
        settings_data = c.fetchone()
        conn.close()
        return settings_data

    def populate_entry_fields(settings_data):
        if settings_data:
            discname1_entry.delete(0, tk.END)
            discname1_entry.insert(0, settings_data[1])
            discount1_entry.delete(0, tk.END)
            discount1_entry.insert(0, settings_data[2])
            discname2_entry.delete(0, tk.END)
            discname2_entry.insert(0, settings_data[3])
            discount2_entry.delete(0, tk.END)
            discount2_entry.insert(0, settings_data[4])
            discname3_entry.delete(0, tk.END)
            discname3_entry.insert(0, settings_data[5])
            discount3_entry.delete(0, tk.END)
            discount3_entry.insert(0, settings_data[6])

    def validate_input():
        # Validation for each input field
        fail = False
        discname1_label.configure(foreground="")
        discname2_label.configure(foreground="")
        discname3_label.configure(foreground="")
        discount1_label.configure(foreground="")
        discount2_label.configure(foreground="")
        discount3_label.configure(foreground="")
        discname1 = discname1_entry.get().strip()
        discname2 = discname2_entry.get().strip()
        discname3 = discname3_entry.get().strip()

        if not discname1 or len(discname1) > 8:
            discname1_label.config(foreground="red")
            fail = True

        if not discname2 or len(discname2) > 8:
            discname2_label.config(foreground="red")
            fail = True

        if not discname3 or len(discname3) > 8:
            discname3_label.config(foreground="red")
            fail = True

        try:
            discount1 = int(discount1_entry.get())
            if not 0 <= discount1 <= 100:
                raise ValueError
        except ValueError:
            discount1_label.config(foreground="red")
            fail = True

        try:
            discount2 = int(discount2_entry.get())
            if not 0 <= discount2 <= 100:
                raise ValueError
        except ValueError:
            discount2_label.config(foreground="red")
            fail = True

        try:
            discount3 = int(discount3_entry.get())
            if not 0 <= discount3 <= 100:
                raise ValueError
        except ValueError:
            discount3_label.config(foreground="red")
            fail = True

        if fail:
            messagebox.showwarning("Missing Required Information", "Please enter valid discount column names (8 letters max) and discount percentages between 0 and 100.", parent=edit_settings_window)
            return False
        return True

    def save_settings():
        if not validate_input():
            return

        # Retrieve values from the entry fields
        discname1 = discname1_entry.get()
        discount1 = int(discount1_entry.get())

        discname2 = discname2_entry.get()
        discount2 = int(discount2_entry.get())

        discname3 = discname3_entry.get()
        discount3 = int(discount3_entry.get())

        # Update the settings in the database
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()

        c.execute("""UPDATE Settings 
                        SET DiscName1 = ?, Discount1 = ?, 
                            DiscName2 = ?, Discount2 = ?, 
                            DiscName3 = ?, Discount3 = ?
                        WHERE Purpose = 'Discounts'""",
                        (discname1, discount1, discname2, discount2, discname3, discount3))
        conn.commit()
        conn.close()

        messagebox.showinfo("Success", "Settings updated successfully.", parent=edit_settings_window)
        edit_settings_window.destroy()

    discname1_label = ttk.Label(edit_settings_window, text="Discount Column Name 1 :")
    discname1_label.grid(row=0, column=0, sticky="e")
    discname1_entry = ttk.Entry(edit_settings_window)
    discname1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    discount1_label = ttk.Label(edit_settings_window, text="Discount 1 Percent :")
    discount1_label.grid(row=1, column=0, sticky="e")
    discount1_entry = ttk.Entry(edit_settings_window)
    discount1_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    discname2_label = ttk.Label(edit_settings_window, text="Discount Column Name 2 :")
    discname2_label.grid(row=2, column=0, sticky="e")
    discname2_entry = ttk.Entry(edit_settings_window)
    discname2_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    discount2_label = ttk.Label(edit_settings_window, text="Discount 2 Percent :")
    discount2_label.grid(row=3, column=0, sticky="e")
    discount2_entry = ttk.Entry(edit_settings_window)
    discount2_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

    discname3_label = ttk.Label(edit_settings_window, text="Discount Column Name 3 :")
    discname3_label.grid(row=4, column=0, sticky="e")
    discname3_entry = ttk.Entry(edit_settings_window)
    discname3_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")

    discount3_label = ttk.Label(edit_settings_window, text="Discount 3 Percent :")
    discount3_label.grid(row=5, column=0, sticky="e")
    discount3_entry = ttk.Entry(edit_settings_window)
    discount3_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")

    save_button = ttk.Button(edit_settings_window, width = 15, text="Apply Changes", command=save_settings)
    save_button.grid(row=6, column=0, pady=10)
    
    cancel_button = ttk.Button(edit_settings_window, width = 15, text="Close", command=edit_settings_window.destroy)
    cancel_button.grid(row=6, column=1, pady=10)
    discname1_entry.focus()

    populate_entry_fields(fetch_settings_data())

def price_list_from_database():
    # Ask for file location and name
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Price_List")
    
    if not file_path:
        # User cancelled file selection
        return
    
    def fetch_settings_data():
        conn = sqlite3.connect("inventory.db")
        c = conn.cursor()
        c.execute("SELECT * FROM Settings WHERE Purpose = 'Discounts'")
        settings_data = c.fetchone()
        conn.close()
        return settings_data

    # Connect to the database
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    
    # Fetch all items from the database
    cursor.execute("""
        SELECT 
            Inventory.Description, 
            Vendors.VendorName, 
            Location.Location, 
            Location.SubLocation,
            Inventory.Sell
        FROM 
            Inventory 
        JOIN 
            Vendors ON Inventory.VendorID = Vendors.VendorID 
        JOIN 
            Location ON Inventory.LocationID = Location.LocationID 
        WHERE 
            Inventory.Quantity > 0
        ORDER BY 
            Inventory.Description, Vendors.VendorName;   
    """)
    data = cursor.fetchall()
    # Close database connection
    conn.close()
    discounts = fetch_settings_data()

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_cols = 'A:G' # the first two cols
    ws.print_title_rows = '1:1' # the first row

    # Write headers
    headers = ["Product Description", "Vendor Name", "Location","Price", discounts[1], discounts[3], discounts[5]]
    ws.append(headers)
    
    # Write data and calculate total cost
    count = 1
    for row in data:
        description, vendor, location, sublocation, sell = row
        count += 1
        ws.append([description, vendor, location + " - " + sublocation, sell / 100, sell * (100 - discounts[2]) / 10000, sell * (100 - discounts[4]) / 10000, sell * (100 - discounts[6]) / 10000])
       
    # Apply currency formatting for Cost, Sell, and Total Cost columns
    currency_format = '[$$-en-US]#,##0.00'
    for col in [4, 5, 6, 7]:  # Columns E, F
        for row in range(2, len(data) + 4):
            ws.cell(row=row, column=col).number_format = currency_format
    
    # Apply bold and underline formatting for the header row
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True, underline='single')

    ws.cell(row=1, column=4).alignment = Alignment(horizontal='right')  
    ws.cell(row=1, column=5).alignment = Alignment(horizontal='right')  
    ws.cell(row=1, column=6).alignment = Alignment(horizontal='right')  
    ws.cell(row=1, column=7).alignment = Alignment(horizontal='right')  
   
    # Center align Vendor, Location, Quantity
    #center_alignment = Alignment(horizontal='center')
    #for col in [2, 3]:  # Columns C, D
    #    for row in range(1, len(data) + 3):
    #        ws.cell(row=row, column=col).alignment = center_alignment
    
    column_widths = {
        'A': 30,    # Width for column A
        'B': 25,    # Width for column B
        'C': 20,    # Width for column C
        'D': 10,     # Width for column D
        'E': 10,    # Width for column E
        'F': 10,     # Width for column F
        'G': 10,     # Width for column G
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save workbook to file
    wb.save(file_path)

def inventory_report_from_database():
    # Ask for file location and name
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Inventory_Cost_Report")
    
    if not file_path:
        # User cancelled file selection
        return
    
    # Connect to the database
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    
    # Fetch all items from the database
    cursor.execute("""
        SELECT 
            Inventory.Description, 
            Vendors.VendorName, 
            Location.Location, 
            Inventory.Quantity, 
            Inventory.Cost
        FROM 
            Inventory 
        JOIN 
            Vendors ON Inventory.VendorID = Vendors.VendorID 
        JOIN 
            Location ON Inventory.LocationID = Location.LocationID 
        WHERE 
            Inventory.Quantity > 0
        ORDER BY 
            Inventory.Description, Vendors.VendorName;   
    """)
    data = cursor.fetchall()
    
    # Close database connection
    conn.close()
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "On Hand Inventory Report"
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_cols = 'A:F' # the first two cols
    ws.print_title_rows = '1:1' # the first row

    # Write headers
    headers = ["Product Description", "Vendor Name", "Location", "On-Hand", "Unit Cost", "Extended"]
    ws.append(headers)
    
    # Write data and calculate total cost
    count = 1
    for row in data:
        description, vendor, location, quantity, cost = row
        count += 1
        ws.append([description, vendor, location, quantity, cost / 100, "=D"+str(count)+"*E"+str(count)])
       
    ws.cell(row=len(data) + 3, column=5).value = "Inventory Total : "
    ws.cell(row=len(data) + 3, column=5).font = Font(bold=True)
    ws.cell(row=len(data) + 3, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=len(data) + 3, column=6).value = "=SUM(F2:F{})".format(len(data) + 1)
    ws.cell(row=len(data) + 3, column=6).font = Font(bold=True)
     
    # Apply currency formatting for Cost, Sell, and Total Cost columns
    currency_format = '[$$-en-US]#,##0.00'
    for col in [5, 6]:  # Columns E, F
        for row in range(2, len(data) + 4):
            ws.cell(row=row, column=col).number_format = currency_format
    
    # Apply bold and underline formatting for the header row
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True, underline='single')

    ws.cell(row=1, column=5).alignment = Alignment(horizontal='right')  
    ws.cell(row=1, column=6).alignment = Alignment(horizontal='right') 

    # Center align Vendor, Location, Quantity
    center_alignment = Alignment(horizontal='center')
    for col in [3, 4]:  # Columns C, D
        for row in range(1, len(data) + 3):
            ws.cell(row=row, column=col).alignment = center_alignment
    
    column_widths = {
        'A': 35,    # Width for column A
        'B': 35,    # Width for column B
        'C': 13,    # Width for column C
        'D': 9,     # Width for column D
        'E': 13,    # Width for column E
        'F': 13,    # Width for column F
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save workbook to file
    wb.save(file_path)

def inventory_count_from_database():
    # Ask for file location and name
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Physical_Count_Sheet")
    
    if not file_path:
        # User cancelled file selection
        return
    
    # Connect to the database
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    
    # Fetch all items from the database
    cursor.execute("""
        SELECT 
            Inventory.Description, 
            Vendors.VendorName, 
            Location.Location,
            Location.SubLocation, 
            Inventory.Quantity 

        FROM 
            Inventory 
        JOIN 
            Vendors ON Inventory.VendorID = Vendors.VendorID 
        JOIN 
            Location ON Inventory.LocationID = Location.LocationID 
        ORDER BY 
            Location.Location, Location.SubLocation, Inventory.Description, Vendors.VendorName;   
    """)
    data = cursor.fetchall()
    
    # Close database connection
    conn.close()
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Count Sheet"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_cols = 'A:F' # the first two cols
    ws.print_title_rows = '1:1' # the first row


    # Write headers
    headers = ["Product Description", "Vendor Name", "Location", "Sub-Location", "On-Hand", "Count", "Diff"]
    ws.append(headers)
    
    # Write data and calculate total cost
    count = 1
    for row in data:
        description, vendor, location, sub, quantity = row
        count += 1
        ws.append([description, vendor, location, sub, quantity, "______", "______"])
       
   # Apply bold and underline formatting for the header row
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True, underline='single')
 
    # Center align Vendor, Location, Quantity, Count, Diff
    center_alignment = Alignment(horizontal='center')
    for col in [3, 4, 5, 6, 7]:  # Columns B, C, D, E, F
        for row in range(1, len(data) + 3):
            ws.cell(row=row, column=col).alignment = center_alignment
    
    column_widths = {
        'A': 30,    # Width for column A
        'B': 30,    # Width for column B
        'C': 13,    # Width for column C
        'D': 13,    # Width for column D
        'E': 9,     # Width for column E
        'F': 9,     # Width for column F
        'G': 9,     # Width for column G
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
 
    # Save workbook to file
    wb.save(file_path)

def open_stats_window():
    def read_stats_data():
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                Inventory.Description, 
                Inventory.Cost,
                Inventory.Sell,
                Vendors.VendorName, 
                Location.Location,
                Inventory.Quantity,
                Inventory.ReorderLevel,
                Inventory.Discontinued       
            FROM 
                Inventory 
            JOIN 
                Vendors ON Inventory.VendorID = Vendors.VendorID 
            JOIN 
                Location ON Inventory.LocationID = Location.LocationID 
        """)
        data = cursor.fetchall()
        conn.close()
        return data

    def calculate_inventory_stats():
        non_discontinued_items = [item for item in inventory_data if item[7] != 'Y']
        total_cost = sum(item[1] * item[5] for item in inventory_data)
        total_sale_value = sum(item[2] * item[5] for item in inventory_data)
        num_products = len(inventory_data)
        num_vendors = len(set(item[3] for item in inventory_data))
        num_locations = len(set(item[4] for item in inventory_data))
        num_items_reorder = sum(1 for item in non_discontinued_items if item[5] <= item[6])
        return total_cost, total_sale_value, num_products, num_vendors, num_locations, num_items_reorder

    inventory_data = read_stats_data()
    total_cost, total_sale_value, num_products, num_vendors, num_locations, num_items_reorder = calculate_inventory_stats()

    stats_window = tk.Toplevel()
    stats_window.title("Inventory Statistics Dashboard")
    center_window(stats_window, 800, 400)
    stats_window.attributes("-topmost", True)  # Keep the window on top
    stats_window.grab_set()   

    stats_window.grid_columnconfigure((0, 1, 2), weight=1, uniform="equal")
    stats_window.rowconfigure(0,weight=1)
    stats_window.rowconfigure(1,weight=1)
    stats_window.rowconfigure(2,weight=0)

    stats_item_frame = tk.Frame(stats_window, bg="white", highlightbackground="black", highlightthickness=1)
    stats_cost_frame = tk.Frame(stats_window, bg="white", highlightbackground="black", highlightthickness=1)
    stats_sell_frame = tk.Frame(stats_window, bg="white", highlightbackground="black", highlightthickness=1)
    stats_vendor_frame = tk.Frame(stats_window, bg="white", highlightbackground="black", highlightthickness=1)
    stats_location_frame = tk.Frame(stats_window, bg="white", highlightbackground="black", highlightthickness=1)
    stats_reorder_frame = tk.Frame(stats_window, bg="white", highlightbackground="black", highlightthickness=1)
 
    stats_item_label = ttk.Label(stats_item_frame, text="# Products", font="bold").pack(pady=10)
    stats_item_value = ttk.Label(stats_item_frame, text=num_products, foreground="blue", font="bold").pack(pady=10)
    stats_cost_label = ttk.Label(stats_cost_frame, text="Total Cost Value", font="bold").pack(pady=10)
    stats_cost_value = ttk.Label(stats_cost_frame, text= "${:,.2f}".format(total_cost/100), foreground="blue", font="bold").pack(pady=10)
    stats_sell_label = ttk.Label(stats_sell_frame, text="Total Sell Value", font="bold").pack(pady=10)
    stats_sell_value = ttk.Label(stats_sell_frame, text="${:,.2f}".format(total_sale_value/100), foreground="blue", font="bold").pack(pady=10)
    stats_vendor_label = ttk.Label(stats_vendor_frame, text="# of Vendors", font="bold").pack(pady=10)
    stats_vendor_value = ttk.Label(stats_vendor_frame, text=num_vendors, foreground="blue", font="bold").pack(pady=10)
    stats_location_label = ttk.Label(stats_location_frame, text="# of Locations", font="bold").pack(pady=10)
    stats_location_value = ttk.Label(stats_location_frame, text=num_locations, foreground="blue", font="bold").pack(pady=10)
    stats_reorder_label = ttk.Label(stats_reorder_frame, text="# to Re-Order", font="bold").pack(pady=10)
    stats_reorder_value = ttk.Label(stats_reorder_frame, text=num_items_reorder, foreground="blue", font="bold").pack(pady=10)

    stats_item_frame.grid(padx=20, pady=20, row=0, column=0, sticky="nesw")
    stats_cost_frame.grid(padx=20, pady=20, row=0, column=1, sticky="nesw")
    stats_sell_frame.grid(padx=20, pady=20, row=0, column=2, sticky="nesw")
    stats_vendor_frame.grid(padx=20, pady=20, row=1, column=0, sticky="nesw")
    stats_location_frame.grid(padx=20, pady=20,row=1, column=1, sticky="nesw")
    stats_reorder_frame.grid(padx=20, pady=20, row=1, column=2, sticky="nesw")

    stats_close_button = ttk.Button(stats_window, text="Close", command=stats_window.destroy)
    stats_close_button.grid(row=2, column=1, pady=20)
    stats_close_button.focus()

root = tk.Tk()
root.title("Hair Salon Product Manager")
center_window(root, 1200, 600)

my_menu = Menu(root)
root.config(menu=my_menu)
menu1 = Menu(my_menu, tearoff=False)
my_menu.add_cascade(label="File", menu=menu1)
menu1.add_command(label="Edit Price Discounts", command=edit_settings_window)
menu1.add_command(label="Create Import Template", command=create_import_file)
menu1.add_command(label="Import Database Records", command=import_records)
menu1.add_command(label="Exit Program", command=lambda: root.destroy())

menu2 = Menu(my_menu, tearoff=False)
my_menu.add_cascade(label="Edit", menu=menu2)
menu2.add_command(label="Edit Vendors", command=open_vendor_window)
menu2.add_command(label="Edit Locations", command=open_locations_window)

menu3 = Menu(my_menu, tearoff=False)
my_menu.add_cascade(label="View", menu=menu3)
menu3.add_command(label="Product Shortages", command=open_shortage_window)
menu3.add_command(label="Inventory Statistics", command=open_stats_window)

menu4 = Menu(my_menu, tearoff=False)
my_menu.add_cascade(label="Reports", menu=menu4)
menu4.add_command(label="Physical Count", command=inventory_count_from_database)
menu4.add_command(label="Inventory Report", command=inventory_report_from_database)
menu4.add_command(label="Price List", command=price_list_from_database)

menu5 = Menu(my_menu, tearoff=False)
my_menu.add_cascade(label="Backup", menu=menu5)
menu5.add_command(label="Backup Database", command=backup_database)
menu5.add_command(label="Restore Database", command=restore_database)

search_frame= tk.Frame(root)
search_label=ttk.Label(search_frame, text="Search for")
search_text = ttk.Entry(search_frame, text="Search")
search_button = ttk.Button(search_frame, text="Search", command=search_treeview)
reset_button = ttk.Button(search_frame, text = "Reset", command=reset_treeview)
sort_frame=tk.Frame(root)
sort_label=ttk.Label(sort_frame, text="Sort by")
sort_text = ttk.Combobox(sort_frame, width = 18, values = ["Product Description", "Vendor Name", "Location", "On Hand", "Reorder", "Delete"])
sort_text.current(0)
sort_text.state(["readonly"])
sort_dn=ttk.Button(sort_frame, text = "A-Z", command = lambda: sort_treeview_up(items, get_mapped_value(sort_text.get())))
sort_up=ttk.Button(sort_frame, text = "Z-A", command = lambda: sort_treeview_dn(items, get_mapped_value(sort_text.get())))
button_frame=tk.Frame(root)
adj_button = ttk.Button(button_frame, text="Transaction", command = open_transaction_window)
edit_button = ttk.Button(button_frame, text = "Edit Product", command=open_edit_item_window)
add_button = ttk.Button(button_frame, text = "Add Product", command=add_item_window)
button_frame_spacer = ttk.Label(button_frame, text = "    ")
inventory_frame = tk.Frame(root)

# Create the Treeview widget
items = ttk.Treeview(inventory_frame, columns=("ItemID", "Description", "Vendor", "Location", "SubLocation", "Quantity", "ReorderLevel", "Cost", "Sell", "Discontinued"))

# Configure column headings

items.heading("#0", text="ItemID")
items.heading("Description", text="Product Description")
items.heading("Vendor", text="Vendor Name")
items.heading("Location", text="Location")
items.heading("SubLocation", text="Sub-Location")
items.heading("Quantity", text="On Hand")
items.heading("ReorderLevel", text="Reorder")
items.heading("Cost", text="Cost")
items.heading("Sell", text="Sell")
items.heading("Discontinued", text="Delete")

# Hide the ItemID column
items["displaycolumns"] = ("Description", "Vendor", "Location", "SubLocation", "Quantity", "ReorderLevel", "Cost", "Sell", "Discontinued")
items.column("#0", width=0, stretch=False)

# Configure individual column settings
# Configure individual column settings with minimum width

items.column("Description", width=250, minwidth=250, anchor="center")  # Description width, centered
items.column("Vendor", width=250, minwidth=250, anchor="center")        # Vendor width, centered
items.column("Location", width=100, minwidth=100, anchor="center")     # Location width, centered
items.column("SubLocation", width=100, minwidth=100, anchor="center")     # SubLocation width, centered
items.column("Quantity", width=55, minwidth=55, anchor="center")      # Quantity width, centered
items.column("ReorderLevel", width=55, minwidth=55, anchor="center")  # Reorder Level width, centered
items.column("Cost", width=80, minwidth=80, anchor="center")           # Cost width, centered
items.column("Sell", width=80, minwidth=80, anchor="center")           # Sell width, centered
items.column("Discontinued", width=20, minwidth=20, anchor="center")  # Discontinued width, centered

# Create vertical scrollbar
v_scrollbar = ttk.Scrollbar(inventory_frame, orient="vertical", command=items.yview)
v_scrollbar.pack(side="right", fill="y")

# Set the scrollbar to control the Treeview
items.configure(yscrollcommand=v_scrollbar.set)


root.columnconfigure(0,weight=1)
root.columnconfigure(1,weight=1)
root.columnconfigure(2,weight=1)
root.rowconfigure(1,weight=1)

search_label.pack(padx=10, pady=5,side="left")
search_text.pack(pady=5, side="left")
search_button.pack(padx=2, pady=5, side="left")
reset_button.pack(padx = 2, pady=5, side="left")
sort_label.pack(padx=10, pady=5,side="left")
sort_text.pack(pady=5, side="left")
sort_dn.pack(padx=1, pady=5, side="left")
sort_up.pack(padx=1, pady=5, side="left")
button_frame_spacer.pack(pady = 5, side="right")
add_button.pack(padx=2, pady=5, side="right")
edit_button.pack(padx=2, pady=5, side="right")
adj_button.pack(padx=2, pady=5, side="right")
search_frame.grid(row=0, column=0, sticky="w")
sort_frame.grid(row=0, column=1)
button_frame.grid(row=0, column=2, sticky="e")

# Pack the Treeview widget
items.pack(side="left", fill="both", expand=True)
inventory_frame.grid(row=1, column=0, columnspan=3, sticky="nsew")

search_text.focus()

check_database()
delete_discontinued_zero_quantity()
delete_discontinued_without_inventory()
delete_vendors_without_inventory()
populate_treeview()
sv_ttk.set_theme("light")
root.mainloop()